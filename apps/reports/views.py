"""WorkSphere HR - Reports Views — Fixed & Complete"""
import csv
import json
from io import BytesIO
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.core.exceptions import PermissionDenied
from django.urls import reverse
from django.utils import timezone
from django.db.models import Sum, Count, Q
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from apps.attendance.models import Attendance
from apps.leave_management.models import LeaveApplication, LeaveBalance
from apps.employees.models import Employee, Department
from drf_spectacular.utils import extend_schema


def _check_reports_access(user):
    if not getattr(user, 'can_access_reports', False):
        raise PermissionDenied('You do not have permission to view reports.')


# ─────────────────────────────────────────────────────────────
# Analytics / Main Dashboard
# ─────────────────────────────────────────────────────────────
@login_required
def reports_dashboard(request):
    _check_reports_access(request.user)
    today = timezone.now().date()
    period_start = today - timezone.timedelta(days=29)

    total_employees = Employee.objects.filter(status='active').count() or 0

    # Aggregate at DB level — never iterate for counting
    att_qs = Attendance.objects.filter(date__gte=period_start, date__lte=today, approval_status='approved')
    agg = att_qs.aggregate(
        present=Count('id', filter=Q(status__in=['present', 'late_mark', 'on_duty'])),
        absent=Count('id', filter=Q(status='absent')),
        leave=Count('id', filter=Q(status='leave')),
        overtime=Sum('overtime_minutes'),
    )
    present_count = agg['present'] or 0
    absent_count = agg['absent'] or 0
    overtime_mins = agg['overtime'] or 0

    # 30-day trend — single aggregated query
    from django.db.models.functions import TruncDate
    daily = (
        att_qs
        .annotate(day=TruncDate('date'))
        .values('day')
        .annotate(
            present=Count('id', filter=Q(status__in=['present', 'late_mark', 'on_duty'])),
            absent=Count('id', filter=Q(status='absent')),
        )
        .order_by('day')
    )
    daily_map = {str(r['day']): r for r in daily}

    trend_labels, trend_present, trend_absent = [], [], []
    for i in range(30):
        d = period_start + timezone.timedelta(days=i)
        key = str(d)
        trend_labels.append(d.strftime('%d %b'))
        rec = daily_map.get(key, {})
        trend_present.append(rec.get('present', 0))
        trend_absent.append(rec.get('absent', 0))

    # Department breakdown
    dept_labels, dept_counts = [], []
    for dept in Department.objects.filter(is_active=True).order_by('name')[:8]:
        cnt = dept.employees.filter(status='active').count()
        if cnt:
            dept_labels.append(dept.name)
            dept_counts.append(cnt)

    # Leave type split
    from apps.leave_management.models import LeaveType
    leave_types = list(LeaveType.objects.values_list('name', flat=True)[:6])
    leave_counts = []
    for lt in LeaveType.objects.all()[:6]:
        leave_counts.append(
            LeaveApplication.objects.filter(leave_type=lt, from_date__gte=period_start, status='approved').count()
        )

    # Top employees table (lightweight)
    top_employees = []
    for emp in Employee.objects.filter(status='active').select_related('department', 'designation')[:15]:
        recs = att_qs.filter(employee=emp)
        total = recs.count() or 1
        present = recs.filter(status__in=['present', 'late_mark', 'on_duty']).count()
        pct = round(present / total * 100, 1)
        ot_hrs = round((recs.aggregate(s=Sum('overtime_minutes'))['s'] or 0) / 60, 1)
        top_employees.append({
            'name': emp.get_full_name(),
            'department': emp.department.name if emp.department else '—',
            'attendance_pct': pct,
            'overtime_hours': ot_hrs,
            'leaves': LeaveApplication.objects.filter(employee=emp, from_date__gte=period_start, status='approved').count(),
            'status': 'Excellent' if pct >= 95 else ('Good' if pct >= 85 else 'Average'),
        })

    context = {
        'today': today,
        'total_employees': total_employees,
        'attendance_rate': round(present_count / max(present_count + absent_count, 1) * 100, 1),
        'leave_utilization': round(agg['leave'] / max(total_employees, 1) * 100, 1),
        'overtime_hours': round(overtime_mins / 60, 1),
        'period_start': period_start,
        'trend_labels': json.dumps(trend_labels),
        'trend_present': json.dumps(trend_present),
        'trend_absent': json.dumps(trend_absent),
        'department_labels': json.dumps(dept_labels),
        'department_scores': json.dumps(dept_counts),
        'leave_labels': json.dumps(leave_types or ['Annual', 'Sick', 'Casual', 'Other']),
        'leave_values': json.dumps(leave_counts or [0, 0, 0, 0]),
        'top_employees': top_employees,
        'departments': Department.objects.filter(is_active=True).order_by('name'),
        'export_csv_url': reverse('reports:attendance_export_csv'),
        'export_excel_url': reverse('reports:attendance_export_excel'),
        'export_pdf_url': reverse('reports:attendance_export_pdf'),
        'active_section': 'analytics',
    }
    return render(request, 'reports/analytics.html', context)


# ─────────────────────────────────────────────────────────────
# Attendance Report Page
# ─────────────────────────────────────────────────────────────
@login_required
def attendance_report(request):
    _check_reports_access(request.user)
    today = timezone.now().date()
    date_from = request.GET.get('date_from', (today - timezone.timedelta(days=29)).isoformat())
    date_to = request.GET.get('date_to', today.isoformat())
    dept_id = request.GET.get('department', '')
    status_filter = request.GET.get('status', '')
    emp_id = request.GET.get('employee_id', '')

    qs = Attendance.objects.filter(approval_status='approved').select_related('employee', 'employee__department', 'employee__designation').order_by('-date', 'employee__first_name')
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if dept_id:
        qs = qs.filter(employee__department_id=dept_id)
    if status_filter:
        qs = qs.filter(status=status_filter)
    if emp_id:
        qs = qs.filter(employee_id=emp_id)

    records = list(qs[:500])
    context = {
        'records': records,
        'total': qs.count(),
        'date_from': date_from,
        'date_to': date_to,
        'dept_id': dept_id,
        'status_filter': status_filter,
        'departments': Department.objects.filter(is_active=True).order_by('name'),
        'employees': Employee.objects.filter(status='active').order_by('first_name', 'last_name'),
        'export_csv_url': reverse('reports:attendance_export_csv'),
        'export_excel_url': reverse('reports:attendance_export_excel'),
        'export_pdf_url': reverse('reports:attendance_export_pdf'),
        'today': today,
        'active_section': 'attendance',
    }
    return render(request, 'reports/attendance_report.html', context)


# ─────────────────────────────────────────────────────────────
# Leave Report Page
# ─────────────────────────────────────────────────────────────
@login_required
def leave_report(request):
    _check_reports_access(request.user)
    today = timezone.now().date()
    date_from = request.GET.get('date_from', today.replace(month=1, day=1).isoformat())
    date_to = request.GET.get('date_to', today.isoformat())
    dept_id = request.GET.get('department', '')
    status_filter = request.GET.get('status', '')

    qs = LeaveApplication.objects.select_related('employee', 'employee__department', 'leave_type').order_by('-from_date')
    if date_from:
        qs = qs.filter(from_date__gte=date_from)
    if date_to:
        qs = qs.filter(from_date__lte=date_to)
    if dept_id:
        qs = qs.filter(employee__department_id=dept_id)
    if status_filter:
        qs = qs.filter(status=status_filter)

    agg = qs.aggregate(
        total=Count('id'),
        approved=Count('id', filter=Q(status='approved')),
        pending=Count('id', filter=Q(status='pending')),
        rejected=Count('id', filter=Q(status='rejected')),
        total_days=Sum('total_days'),
    )
    records = list(qs[:500])
    context = {
        'records': records,
        'agg': agg,
        'date_from': date_from,
        'date_to': date_to,
        'dept_id': dept_id,
        'status_filter': status_filter,
        'departments': Department.objects.filter(is_active=True).order_by('name'),
        'export_csv_url': reverse('reports:leave_export_csv'),
        'export_excel_url': reverse('reports:leave_export_excel'),
        'today': today,
        'active_section': 'leave',
    }
    return render(request, 'reports/leave_report.html', context)


# ─────────────────────────────────────────────────────────────
# Employee Report Page
# ─────────────────────────────────────────────────────────────
@login_required
def employee_report(request):
    _check_reports_access(request.user)
    dept_id = request.GET.get('department', '')
    status_filter = request.GET.get('status', 'active')

    qs = Employee.objects.select_related('department', 'designation', 'location').order_by('first_name', 'last_name')
    if dept_id:
        qs = qs.filter(department_id=dept_id)
    if status_filter:
        qs = qs.filter(status=status_filter)

    context = {
        'employees': list(qs[:500]),
        'total': qs.count(),
        'dept_id': dept_id,
        'status_filter': status_filter,
        'departments': Department.objects.filter(is_active=True).order_by('name'),
        'export_csv_url': reverse('reports:employee_export_csv'),
        'export_excel_url': reverse('reports:employee_export_excel'),
        'active_section': 'employees',
    }
    return render(request, 'reports/employee_report.html', context)


# ─────────────────────────────────────────────────────────────
# Payroll Report Page
# ─────────────────────────────────────────────────────────────
@login_required
def payroll_report(request):
    if not getattr(request.user, 'can_manage_payroll', False) and not getattr(request.user, 'can_access_reports', False):
        raise PermissionDenied()
    today = timezone.now().date()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    dept_id = request.GET.get('department', '')

    try:
        from apps.payroll.models import PayrollPeriod, PayslipRecord
        period_qs = PayrollPeriod.objects.filter(year=year, month=month).first()
        if period_qs:
            qs = PayslipRecord.objects.filter(payroll_period=period_qs).select_related('employee', 'employee__department', 'payroll_period')
            if dept_id:
                qs = qs.filter(employee__department_id=dept_id)
            agg = qs.aggregate(total_gross=Sum('gross_earnings'), total_net=Sum('net_salary'), total_deductions=Sum('total_deductions'), count=Count('id'))
            payslips = list(qs[:500])
        else:
            qs = PayslipRecord.objects.none()
            agg = {}
            payslips = []
    except Exception:
        payslips, agg = [], {}

    context = {
        'payslips': payslips,
        'agg': agg,
        'month': month,
        'year': year,
        'dept_id': dept_id,
        'departments': Department.objects.filter(is_active=True).order_by('name'),
        'months': [(i, ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][i-1]) for i in range(1, 13)],
        'years': list(range(today.year - 3, today.year + 1)),
        'active_section': 'payroll',
    }
    return render(request, 'reports/payroll_report.html', context)


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────
def _build_att_qs(request):
    today = timezone.now().date()
    date_from = request.GET.get('date_from', (today - timezone.timedelta(days=29)).isoformat())
    date_to = request.GET.get('date_to', today.isoformat())
    qs = Attendance.objects.filter(approval_status='approved').select_related('employee').order_by('-date')
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if request.GET.get('status'):
        qs = qs.filter(status=request.GET['status'])
    if request.GET.get('department'):
        qs = qs.filter(employee__department_id=request.GET['department'])
    if request.GET.get('employee_id'):
        qs = qs.filter(employee_id=request.GET['employee_id'])
    return qs


def _build_leave_qs(request):
    today = timezone.now().date()
    date_from = request.GET.get('date_from', today.replace(month=1, day=1).isoformat())
    date_to = request.GET.get('date_to', today.isoformat())
    qs = LeaveApplication.objects.select_related('employee', 'leave_type').order_by('-from_date')
    if date_from:
        qs = qs.filter(from_date__gte=date_from)
    if date_to:
        qs = qs.filter(from_date__lte=date_to)
    if request.GET.get('status'):
        qs = qs.filter(status=request.GET['status'])
    if request.GET.get('department'):
        qs = qs.filter(employee__department_id=request.GET['department'])
    return qs


# ─────────────────────────────────────────────────────────────
# Attendance Exports
# ─────────────────────────────────────────────────────────────
@login_required
def attendance_report_export_csv(request):
    _check_reports_access(request.user)
    qs = _build_att_qs(request)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.csv"'
    writer = csv.writer(response)
    writer.writerow(['Employee', 'Date', 'Check In', 'Check Out', 'Status', 'Working Minutes', 'Overtime Minutes'])
    for r in qs.iterator():
        writer.writerow([
            r.employee.get_full_name(), r.date.isoformat(),
            r.clock_in.strftime('%H:%M') if r.clock_in else '',
            r.clock_out.strftime('%H:%M') if r.clock_out else '',
            r.get_status_display(), r.effective_working_minutes, r.overtime_minutes,
        ])
    return response


@login_required
def attendance_report_export_excel(request):
    _check_reports_access(request.user)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    qs = _build_att_qs(request)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'
    headers = ['Employee', 'Date', 'Check In', 'Check Out', 'Status', 'Working Minutes', 'Overtime Minutes']
    ws.append(headers)
    hdr_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
    for r in qs.iterator():
        ws.append([
            r.employee.get_full_name(), r.date.isoformat(),
            r.clock_in.strftime('%H:%M') if r.clock_in else '',
            r.clock_out.strftime('%H:%M') if r.clock_out else '',
            r.get_status_display(), r.effective_working_minutes, r.overtime_minutes,
        ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(stream.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.xlsx"'
    return response


@login_required
def attendance_report_export_pdf(request):
    _check_reports_access(request.user)
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    qs = _build_att_qs(request)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=18, leftMargin=18, topMargin=18, bottomMargin=18)
    styles = getSampleStyleSheet()
    elements = [Paragraph('Attendance Report — WorkSphere HR', styles['Heading2']), Spacer(1, 10)]
    data = [['Employee', 'Date', 'Check In', 'Check Out', 'Status', 'Working Min', 'OT Min']]
    for r in qs[:500]:
        data.append([
            r.employee.get_full_name(), str(r.date),
            r.clock_in.strftime('%H:%M') if r.clock_in else '—',
            r.clock_out.strftime('%H:%M') if r.clock_out else '—',
            r.get_status_display(), str(r.effective_working_minutes), str(r.overtime_minutes),
        ])
    table = Table(data, repeatRows=1, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('ALIGN',      (0, 0), (-1, -1), 'LEFT'),
        ('GRID',       (0, 0), (-1, -1), 0.25, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    elements.append(table)
    doc.build(elements)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.pdf"'
    return response


# ─────────────────────────────────────────────────────────────
# Leave Exports
# ─────────────────────────────────────────────────────────────
@login_required
def leave_report_export_csv(request):
    _check_reports_access(request.user)
    qs = _build_leave_qs(request)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="leave_report.csv"'
    writer = csv.writer(response)
    writer.writerow(['Employee', 'Leave Type', 'From Date', 'To Date', 'Days', 'Status', 'Reason'])
    for r in qs.iterator():
        writer.writerow([
            r.employee.get_full_name(), r.leave_type.name,
            r.from_date.isoformat(), r.to_date.isoformat(),
            getattr(r, 'number_of_days', ''), r.status, r.reason or '',
        ])
    return response


@login_required
def leave_report_export_excel(request):
    _check_reports_access(request.user)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    qs = _build_leave_qs(request)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Leave Report'
    headers = ['Employee', 'Leave Type', 'From Date', 'To Date', 'Days', 'Status', 'Reason']
    ws.append(headers)
    fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
    for r in qs.iterator():
        ws.append([
            r.employee.get_full_name(), r.leave_type.name,
            r.from_date.isoformat(), r.to_date.isoformat(),
            getattr(r, 'number_of_days', ''), r.status, r.reason or '',
        ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    resp = HttpResponse(stream.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="leave_report.xlsx"'
    return resp


# ─────────────────────────────────────────────────────────────
# Employee Exports
# ─────────────────────────────────────────────────────────────
@login_required
def employee_report_export_csv(request):
    _check_reports_access(request.user)
    dept_id = request.GET.get('department', '')
    qs = Employee.objects.select_related('department', 'designation').order_by('first_name')
    if dept_id:
        qs = qs.filter(department_id=dept_id)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="employee_report.csv"'
    writer = csv.writer(response)
    writer.writerow(['Employee ID', 'Name', 'Email', 'Department', 'Designation', 'Status', 'Date of Joining'])
    for emp in qs.iterator():
        writer.writerow([
            emp.employee_id, emp.get_full_name(), emp.user.email if emp.user else '',
            emp.department.name if emp.department else '',
            emp.designation.name if emp.designation else '',
            emp.status, emp.date_of_joining.isoformat() if emp.date_of_joining else '',
        ])
    return response


@login_required
def employee_report_export_excel(request):
    _check_reports_access(request.user)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    dept_id = request.GET.get('department', '')
    qs = Employee.objects.select_related('department', 'designation').order_by('first_name')
    if dept_id:
        qs = qs.filter(department_id=dept_id)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Employee Report'
    headers = ['Employee ID', 'Name', 'Email', 'Department', 'Designation', 'Status', 'Date of Joining']
    ws.append(headers)
    fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
    for emp in qs.iterator():
        ws.append([
            emp.employee_id, emp.get_full_name(), emp.user.email if emp.user else '',
            emp.department.name if emp.department else '',
            emp.designation.name if emp.designation else '',
            emp.status, emp.date_of_joining.isoformat() if emp.date_of_joining else '',
        ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    resp = HttpResponse(stream.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="employee_report.xlsx"'
    return resp


# ─────────────────────────────────────────────────────────────
# API Views
# ─────────────────────────────────────────────────────────────
@extend_schema(request=None, responses=None)
class ReportsSummaryAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        _check_reports_access(request.user)
        today = timezone.now().date()
        period_start = today - timezone.timedelta(days=30)
        return Response({
            'employee_count': Employee.objects.filter(status='active').count(),
            'attendance_count': Attendance.objects.filter(date__gte=period_start).count(),
            'pending_leaves': LeaveApplication.objects.filter(status='pending').count(),
            'reporting_period': {'start': period_start.isoformat(), 'end': today.isoformat()},
        })


@extend_schema(request=None, responses=None)
class AttendanceReportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        _check_reports_access(request.user)
        qs = _build_att_qs(request)
        data = [
            {
                'employee': r.employee.get_full_name(),
                'date': r.date.isoformat(),
                'status': r.status,
                'clock_in': r.clock_in.strftime('%H:%M') if r.clock_in else None,
                'clock_out': r.clock_out.strftime('%H:%M') if r.clock_out else None,
                'overtime_minutes': r.overtime_minutes,
            }
            for r in qs[:200]
        ]
        return Response({'count': qs.count(), 'results': data})
